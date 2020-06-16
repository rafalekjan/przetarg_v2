from selenium import webdriver
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook, load_workbook


def otworz_strone_selenium():
    strona = webdriver.Chrome('C:\Python27\Scripts\chromedriver.exe')
    strona.implicitly_wait(10)
    strona.get(aktualna_strona_www)
    return strona


def otworz_strone_soup():
    strona = requests.get(aktualna_strona_www)
    dane = BeautifulSoup(strona.content, 'html.parser')
    return dane


def ilosc_obiektow_selenium(strona):
    return len(strona.find_elements_by_css_selector(aktualna_css_strona))


def ilosc_obiektow_soup(strona):
    return len(strona.select(aktualna_css_strona))


def tytul_selenium(strona, numer_obiektu):
    css_nth = aktualna_css_strona + ":nth-of-type(" + str(numer_obiektu + 1) + ")"
    if len(aktualna_css_tytul) > 0:
        return strona.find_element_by_css_selector(css_nth + ' > ' + aktualna_css_tytul).text.lower()
    return strona.find_element_by_css_selector(css_nth).text.lower()


def tytul_soup(strona, numer_obiektu):
    css_nth = aktualna_css_strona + ":nth-of-type(" + str(numer_obiektu + 1) + ")"
    if len(aktualna_css_tytul) > 0:
        return strona.select(css_nth)[0].select(aktualna_css_tytul)[0].text.lower()
    return strona.select(css_nth)[0].text.lower()


def adres_selenium(strona, numer_obiektu):
    css_nth = aktualna_css_strona + ":nth-of-type(" + str(numer_obiektu + 1) + ")"
    if len(aktualna_css_adres) > 0:
        if aktualna_podst_adres is not None:
            return aktualna_podst_adres + strona.find_element_by_css_selector(css_nth + ' > ' + css_adresy).get_attribute('href')
        else:
            return strona.find_element_by_css_selector(css_nth + ' > ' + css_adresy).get_attribute('href')
    else:
        return aktualna_podst_adres + strona.find_element_by_css_selector(css_nth).get_attribute('href')


def adres_soup(strona, numer_obiektu):
    css_nth = aktualna_css_strona + ":nth-of-type(" + str(numer_obiektu + 1) + ")"
    if len(aktualna_css_adres) > 0:
        if aktualna_podst_adres is not None:
            return aktualna_podst_adres + strona.select(css_nth)[0].select(aktualna_css_adres, href=True)[0]['href']
        else:
            return strona.select(css_nth)[0].select(aktualna_css_adres, href=True)[0]['href']
    else:
        return aktualna_podst_adres


def adres_trovit(strona, numer_obiektu):
    css_nth = aktualna_css_strona + ":nth-of-type(" + str(numer_obiektu + 1) + ")"
    id_ogloszenia = strona.select(css_nth)[0].select(aktualna_css_adres, href=True)[0]['data-id']
    strona_trovit = "http://rd.clk.thribee.com/id." + id_ogloszenia + "/origin.2/section.1/section_type.1/country.pl/vertical.homes/"
    return strona_trovit


def odczyt_selenium():
    dane_strony = otworz_strone_selenium()
    ilosc_obiektow = ilosc_obiektow_selenium(dane_strony)
    for nr in range(ilosc_obiektow):
        lista_tytulow.append(tytul_selenium(dane_strony, nr))
        lista_adresow.append(adres_selenium(dane_strony, nr))


def odczyt_soup():
    dane_strony = otworz_strone_soup()
    ilosc_obiektow = ilosc_obiektow_soup(dane_strony)
    for nr in range(ilosc_obiektow):
        try:
            lista_tytulow.append(tytul_soup(dane_strony, nr))
        except:
            # print("Blad tytulu na nr. ", nr)
            lista_tytulow.append("Blad tytulu")
        try:
            if aktualna_metoda_www == "trovit":
                lista_adresow.append(adres_trovit(dane_strony, nr))
            else:
                lista_adresow.append(adres_soup(dane_strony, nr))
        except:
            # print("Blad adresu na nr. ", nr)
            lista_adresow.append("Blad adresu")


def odczyt_historii():
    historia_adresow = odczyt_danych_kolumna("historia.xlsx", 0, 10)
    return historia_adresow


def zapis_exela(nr, nazwa):
    plik = load_workbook(nazwa + ".xlsx")
    wyniki = plik.active
    if len(lista_adresow[nr]) > 255:
        wyniki.cell(row=len(wyniki["A"]) + 1, column=1).value = '=HYPERLINK("{}", "{}")'.format(aktualna_strona_www, ">Link<")
    else:
        wyniki.cell(row=len(wyniki["A"]) + 1, column=1).value = '=HYPERLINK("{}", "{}")'.format(lista_adresow[nr], ">Link<")
    wyniki.cell(row=len(wyniki["A"]), column=2).value = lista_tytulow[nr]
    wyniki.cell(row=len(wyniki["A"]), column=10).value = lista_adresow[nr]
    plik.save(nazwa + ".xlsx")


def koniec_zapisu_exela(nazwa):
    plik = load_workbook(nazwa + ".xlsx")
    wyniki = plik.active
    wyniki.cell(row=len(wyniki["A"]) + 1, column=1).value = 'xxx'
    wyniki.cell(row=len(wyniki["A"]), column=2).value = 'xxx'
    wyniki.cell(row=len(wyniki["A"]), column=10).value = 'xxx'
    plik.save(nazwa + ".xlsx")


def tworzenie_raportu():
    try:
        load_workbook("Raport_" + aktualna_data + ".xlsx")
    except:
        plik_raportu = Workbook()
        plik_raportu.save("Raport_" + aktualna_data + ".xlsx")


def tworzenie_historii():
    try:
        load_workbook("historia.xlsx")
    except:
        plik_historii = Workbook()
        plik_historii.save("historia.xlsx")


def sprawdzenie_historii():
    for nr, adres in enumerate(lista_adresow):
        if adres not in aktualna_historia:
            zapis_exela(nr, "historia")
            tworzenie_raportu()
            zapis_exela(nr, "Raport_" + aktualna_data)


def start_programu():
    if aktualna_metoda_www == "selenium":
        odczyt_selenium()
        sprawdzenie_historii()
    if aktualna_metoda_www == "soup" or aktualna_metoda_www == "trovit":
        odczyt_soup()
        sprawdzenie_historii()


def odczytaj_z_excela(excel, kol, rza):
    return excel.cell(row=kol, column=rza).value


def odczyt_danych_kolumna(plik_excel, strona, kolumna):
    dane = []
    otwarty_excel = load_workbook(plik_excel)
    otwarty_excel_lista = load_workbook(plik_excel).sheetnames
    #
    otwarty_excel_tekst = otwarty_excel[otwarty_excel_lista[strona]]
    #
    for i in range(1, len(otwarty_excel_tekst["A"]) + 1):
        dane.append(odczytaj_z_excela(otwarty_excel_tekst, i, kolumna))
    return dane


def sprawdz_dane_excela():
    if not len(strony_www) == len(css_strony) == len(css_tytuly) == len(css_adresy) == len(podst_adresy) == len(metody_www):
        breakpoint()


####################
####################

strony_www = odczyt_danych_kolumna("Baza_stron.xlsx", 1, 4)
css_strony = odczyt_danych_kolumna("Baza_stron.xlsx", 1, 5)
css_tytuly = odczyt_danych_kolumna("Baza_stron.xlsx", 1, 8)
css_adresy = odczyt_danych_kolumna("Baza_stron.xlsx", 1, 7)
podst_adresy = odczyt_danych_kolumna("Baza_stron.xlsx", 1, 9)
metody_www = odczyt_danych_kolumna("Baza_stron.xlsx", 1, 6)
sprawdz_dane_excela()
tworzenie_historii()

aktualna_data = time.strftime("%Y_%m_%d")
lista_tytulow = []
lista_adresow = []

aktualnych_historycznych = len(odczyt_historii())

for i in range(1, 34):
    aktualna_historia = odczyt_historii()
    aktualna_strona_www = strony_www[i]
    aktualna_css_strona = css_strony[i]
    aktualna_css_tytul = css_tytuly[i]
    aktualna_css_adres = css_adresy[i]
    aktualna_podst_adres = podst_adresy[i]
    aktualna_metoda_www = metody_www[i]
    start_programu()

koniec_zapisu_exela("Raport_" + aktualna_data)
aktualna_historia = odczyt_historii()
print("Nowych wpisów w historii :", len(aktualna_historia) - aktualnych_historycznych)



